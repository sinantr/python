#!/usr/bin/env python
# -*- coding: utf-8 -*-

import defFtp, defMySql, sys
from openpyxl import load_workbook
import defMandrill


filename = 'data.xlsx'

# dosyayı indir
defFtp.ftpBaglanti(filename)
filename = '/home/python/xxxxxxxxxx/fiyat_guncel/' + filename
try:
	# exceli açıyoruz
	wb = load_workbook(filename = '/home/python/xxxxxxxxxx/fiyat_guncel/data.xlsx', use_iterators = True)   
	sheet = wb.get_active_sheet() 
	
	# dataları bir listenin (list) içinde toplayacağız 
	table=[]  	
	for row in sheet.iter_rows():
		line=[]
		for cell in row:
			line.append(cell.internal_value)
		table.append(line)			
except:
	a = 'Dosya açılamadı /home/python/xxxxxxxxxx/%s' % filename
	print a
	#defMandrill.defMandrillRoot(a,a)
	sys.exit()


print 'Update başladı.'

for row in table:
	try:
		defMySql.baglanti(int(row[0]),row[1],int(row[2]))
		print "%s - %s - %s" %(int(row[0]),row[1], int(row[2]))	
	except:
		pass
	
sonuc = '%s satırdan uygun olanlar update edildi.' %(len(table)) 
print sonuc

def mailAt():
	txt = ''
	for row in table:	
		txt += "%s - %s TL - %s<br>" % (int(row[0]),row[1],int(row[2]))
	return txt	

html = """%s<br>
-----------------------------------------------------------------<p>
%s<p>
Saygılarımızla.<p>
<p><p><hr>
<small>
<b>Açıklama</b><br>
Bu rapor otomatik olarak oluşturulmuştur.<br>
xxxxxxxxxx.com sitesindeki dataların güncellenmesi 1 (bir) saat içinde gerçekleşir.
</small>
""" %(sonuc,mailAt())



defMandrill.defMandrillUser(html,sonuc)   
defMandrill.defMandrillRoot(html,sonuc)






